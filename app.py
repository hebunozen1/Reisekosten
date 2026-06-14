import os
import sqlite3
import uuid
import hashlib
import smtplib
import io
import csv
import zipfile
import tempfile
import shutil
import posixpath
import re
import time
from datetime import datetime, timedelta, timezone
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from flask import (
    Flask, render_template, request, redirect,
    url_for, session, flash, send_from_directory, send_file, jsonify
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge

import requests
import psycopg2
import psycopg2.extras

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =====================
# Flask App
# =====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("SQLITE_DB_PATH", os.path.join(BASE_DIR, "reisekosten.db"))

# =====================
# Kategorien (Keys -> Sprachen)
# =====================
CATEGORIES = {
    "meal": {"ar": "وجبات", "de": "Verpflegung"},
    "water": {"ar": "مياه الشرب", "de": "Trinkwasser"},
    "hotel": {"ar": "فندق", "de": "Hotel"},
    "transport": {"ar": "مواصلات", "de": "Transport"},
    "shopping": {"ar": "تسوق", "de": "Einkauf"},
    "other": {"ar": "أخرى", "de": "Sonstiges"},
    "taxi": {"ar": "تاكسي", "de": "Taxi"},
    "tip_bus": {"ar": "إكرامية سائق الحافلة", "de": "Trinkgeld Busfahrer"},
    "tip_hotel": {"ar": "إكرامية موظفي الفندق", "de": "Trinkgeld Hotelmitarbeiter"},
}


def resolve_category(value: str | None) -> tuple[str, str]:
    """Return Arabic/German category labels for both new and older form values.

    Older deployments or cached pages sometimes submit the visible Arabic label
    instead of the internal category key. Without this fallback, kategorie_de can
    become NULL and the database insert fails with an Internal Server Error.
    """
    raw = (value or "").strip()

    if raw in CATEGORIES:
        category = CATEGORIES[raw]
        return category["ar"], category["de"]

    for category in CATEGORIES.values():
        if raw in (category["ar"], category["de"]):
            return category["ar"], category["de"]

    if raw:
        return raw, CATEGORIES["other"]["de"]

    return CATEGORIES["other"]["ar"], CATEGORIES["other"]["de"]
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret")


# Uploads
app.config["UPLOAD_FOLDER"] = os.path.join(BASE_DIR, "uploads")
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# Upload-Limit: Backups mit vielen Belegen können deutlich größer als einzelne Handyfotos sein.
# In Render kann der Wert bei Bedarf über MAX_UPLOAD_MB angepasst werden.
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_UPLOAD_MB", "300")) * 1024 * 1024


@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(error):
    flash("Die hochgeladene Datei ist zu groß. Bitte ein kleineres Bild/Backup hochladen oder MAX_UPLOAD_MB in Render erhöhen.", "err")
    return redirect(request.referrer or url_for("dashboard"))


def save_uploaded_file(file_storage, prefix="beleg"):
    """Save an uploaded receipt image safely and return the stored file name.

    The function creates the upload folder right before saving, handles filenames
    with Arabic/special characters and raises a clear error instead of causing
    a generic internal server error.
    """
    if not file_storage or not getattr(file_storage, "filename", ""):
        return None

    upload_dir = app.config["UPLOAD_FOLDER"]
    os.makedirs(upload_dir, exist_ok=True)

    original_name = secure_filename(file_storage.filename or "")
    if not original_name:
        extension = ""
        mimetype = (getattr(file_storage, "mimetype", "") or "").lower()
        if "jpeg" in mimetype or "jpg" in mimetype:
            extension = ".jpg"
        elif "png" in mimetype:
            extension = ".png"
        elif "webp" in mimetype:
            extension = ".webp"
        elif "pdf" in mimetype:
            extension = ".pdf"
        else:
            extension = ".jpg"
        original_name = f"{prefix}{extension}"

    stored_name = f"{uuid.uuid4().hex}_{original_name}"
    target_path = os.path.join(upload_dir, stored_name)

    try:
        file_storage.save(target_path)
    except Exception as exc:
        app.logger.exception("Beleg-Upload konnte nicht gespeichert werden: %s", original_name)
        raise RuntimeError(f"Beleg konnte nicht gespeichert werden: {exc}") from exc

    return stored_name


def get_first_uploaded_file(*field_names):
    """Return the first uploaded file that actually has a filename.

    The current form uses one upload field so mobile browsers can offer
    camera, gallery, or file picker automatically. Older split camera/gallery
    field names remain supported for already deployed templates.
    """
    for field_name in field_names:
        for file_storage in request.files.getlist(field_name):
            if file_storage and getattr(file_storage, "filename", ""):
                return file_storage
    return None


# =====================
# Database
# =====================
DATABASE_URL = os.environ.get("DATABASE_URL")
APPROVER_EMAIL = (os.environ.get("APPROVAL_APPROVER_EMAIL") or "hebun.oezen@gmx.de").strip().lower()

ARABIC_RE = re.compile(r"[\u0600-\u06FF]")


def contains_arabic(text: str | None) -> bool:
    """Return True if a text contains Arabic characters."""
    return bool(ARABIC_RE.search(text or ""))


def translation_is_missing_or_not_german(source_text: str | None, translated_text: str | None) -> bool:
    """Detect old entries where beschreibung_de is empty or still contains the Arabic original."""
    source = (source_text or "").strip()
    translated = (translated_text or "").strip()
    if not source or not contains_arabic(source):
        return False
    return (not translated) or translated == source or contains_arabic(translated)


def translate_arabic_to_german(text: str | None) -> str:
    """Translate Arabic expense descriptions to German. Fallback keeps the original text."""
    source = (text or "").strip()
    if not source:
        return ""
    if not contains_arabic(source):
        return source

    if not _env_bool("AUTO_TRANSLATE_DESCRIPTIONS", True):
        return source

    try:
        timeout_seconds = float(os.environ.get("TRANSLATE_TIMEOUT_SECONDS", "2.5"))
        response = requests.get(
            "https://translate.googleapis.com/translate_a/single",
            params={
                "client": "gtx",
                "sl": "ar",
                "tl": "de",
                "dt": "t",
                "q": source,
            },
            timeout=timeout_seconds,
        )
        response.raise_for_status()
        data = response.json()
        translated = "".join(part[0] for part in data[0] if part and part[0]).strip()
        return translated or source
    except Exception as exc:
        app.logger.warning("Beschreibung konnte nicht automatisch übersetzt werden: %s", exc)
        return source


def translate_and_store_missing_descriptions(cur, kosten_rows, max_items=8, time_budget_seconds=10):
    """Translate a small batch of missing German descriptions and persist them.

    Important: this function is intentionally limited. External translation calls can be slow
    or unavailable. A bulk translation during normal dashboard loading can block the Render
    worker and cause a timeout.
    """
    changed = 0
    started = time.monotonic()
    for row in kosten_rows:
        if max_items is not None and changed >= max_items:
            break
        if time_budget_seconds is not None and (time.monotonic() - started) >= time_budget_seconds:
            break

        source = row.get("beschreibung_ar")
        current = row.get("beschreibung_de")
        if not translation_is_missing_or_not_german(source, current):
            continue

        translated = translate_arabic_to_german(source)
        if not translated or translated.strip() == (source or "").strip():
            continue

        row["beschreibung_de"] = translated
        cur.execute(
            "UPDATE kosten SET beschreibung_de=%s WHERE id=%s" if DATABASE_URL else
            "UPDATE kosten SET beschreibung_de=? WHERE id=?",
            (translated, row.get("id")),
        )
        changed += 1
    return changed

def normalize_email(email: str | None) -> str:
    """Normalize emails for consistent lookup (trim + lower)."""
    return (email or "").strip().lower()


def row_value(row, key, default=None):
    """Return a column value for both dict-like rows and sqlite3.Row."""
    if row is None:
        return default
    if hasattr(row, "keys") and key in row.keys():
        return row[key]
    try:
        return row[key]
    except Exception:
        return default


def _as_utc_naive(value):
    """Convert datetimes to naive UTC so SQLite and PostgreSQL values are comparable."""
    if value is None:
        return None
    if value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def parse_reset_expiry(raw_value):
    """Parse reset expiry values from SQLite text or native datetime objects."""
    if raw_value is None:
        return None
    if isinstance(raw_value, datetime):
        return _as_utc_naive(raw_value)
    if isinstance(raw_value, str):
        value = raw_value.strip()
        if not value:
            return None
        try:
            return _as_utc_naive(datetime.fromisoformat(value))
        except ValueError:
            pass
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None

def get_db():
    if DATABASE_URL:
        conn = psycopg2.connect(
            DATABASE_URL,
            connect_timeout=5,
            cursor_factory=psycopg2.extras.RealDictCursor,
        )
        return conn
    else:
        conn = sqlite3.connect(
            DB_PATH,
            timeout=10,
            check_same_thread=False
        )
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        return conn


def ensure_user_reset_columns(cur):
    if DATABASE_URL:
        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = 'users'
              AND column_name IN ('reset_token', 'reset_token_expires')
        """)
        existing = {row['column_name'] if hasattr(row, 'keys') else row[0] for row in cur.fetchall()}
        if 'reset_token' not in existing:
            cur.execute("ALTER TABLE users ADD COLUMN reset_token TEXT")
        if 'reset_token_expires' not in existing:
            cur.execute("ALTER TABLE users ADD COLUMN reset_token_expires TIMESTAMP")
    else:
        cur.execute("PRAGMA table_info(users)")
        existing = {row['name'] if hasattr(row, 'keys') else row[1] for row in cur.fetchall()}
        if 'reset_token' not in existing:
            cur.execute("ALTER TABLE users ADD COLUMN reset_token TEXT")
        if 'reset_token_expires' not in existing:
            cur.execute("ALTER TABLE users ADD COLUMN reset_token_expires TEXT")

def ensure_user_approval_columns(cur):
    ensure_column(
        cur,
        "users",
        "approval_status",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS approval_status TEXT NOT NULL DEFAULT 'approved'",
        "ALTER TABLE users ADD COLUMN approval_status TEXT NOT NULL DEFAULT 'approved'",
    )
    ensure_column(
        cur,
        "users",
        "approved_at",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS approved_at TIMESTAMP",
        "ALTER TABLE users ADD COLUMN approved_at TEXT",
    )
    ensure_column(
        cur,
        "users",
        "approved_by_email",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS approved_by_email TEXT",
        "ALTER TABLE users ADD COLUMN approved_by_email TEXT",
    )


def ensure_column(cur, table_name, column_name, postgres_sql, sqlite_sql):
    if DATABASE_URL:
        cur.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = %s AND column_name = %s
            """,
            (table_name, column_name),
        )
        exists = cur.fetchone()
        if not exists:
            cur.execute(postgres_sql)
    else:
        cur.execute(f"PRAGMA table_info({table_name})")
        columns = {row['name'] if hasattr(row, 'keys') else row[1] for row in cur.fetchall()}
        if column_name not in columns:
            cur.execute(sqlite_sql)


def ensure_schema():
    conn = get_db()
    cur = conn.cursor()

    # ---- users ----
    if DATABASE_URL:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT,
            email TEXT UNIQUE,
            password_hash TEXT,
            role TEXT,
            reset_token TEXT,
            reset_token_expires TIMESTAMP
        )
        """)
    else:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            email TEXT UNIQUE,
            password_hash TEXT,
            role TEXT,
            reset_token TEXT,
            reset_token_expires TEXT
        )
        """)

    ensure_user_reset_columns(cur)
    ensure_user_approval_columns(cur)

    # ---- startguthaben ----
    if DATABASE_URL:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS startguthaben (
            user_id INTEGER PRIMARY KEY REFERENCES users(id) ON DELETE CASCADE,
            betrag_sar NUMERIC NOT NULL DEFAULT 0,
            wechselkurs_beleg TEXT
        )
        """)
    else:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS startguthaben (
            user_id INTEGER PRIMARY KEY,
            betrag_sar REAL NOT NULL DEFAULT 0,
            wechselkurs_beleg TEXT
        )
        """)

    # ---- kosten ----
    if DATABASE_URL:
        cur.execute("""
    CREATE TABLE IF NOT EXISTS kosten (
        id BIGSERIAL PRIMARY KEY,
        user_id INTEGER,
        datum DATE,
        kategorie_ar TEXT,
        kategorie_de TEXT NOT NULL DEFAULT '',
        beschreibung_ar TEXT,
        beschreibung_de TEXT NOT NULL DEFAULT '',
        betrag_sar NUMERIC NOT NULL DEFAULT 0,
        beleg TEXT,
        ohne_beleg INTEGER NOT NULL DEFAULT 0,
        genehmigt INTEGER NOT NULL DEFAULT 0,
        genehmigt_von TEXT,
        created_at TIMESTAMP DEFAULT NOW(),
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    )
""")
    else:
        cur.execute("""
    CREATE TABLE IF NOT EXISTS kosten (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
        datum TEXT,
        kategorie_ar TEXT,
        kategorie_de TEXT NOT NULL DEFAULT '',
        beschreibung_ar TEXT,
        beschreibung_de TEXT NOT NULL DEFAULT '',
        betrag_sar REAL NOT NULL DEFAULT 0,
        beleg TEXT,
        ohne_beleg INTEGER NOT NULL DEFAULT 0,
        genehmigt INTEGER NOT NULL DEFAULT 0,
        genehmigt_von TEXT,
        created_at TEXT
    )
""")


    # ---- kosten migrations for existing databases ----
    # Ältere Datenbanken wurden teilweise ohne diese Spalten angelegt.
    # Ohne diese Migration stürzt das Speichern einer Ausgabe mit Beleg beim INSERT ab.
    ensure_column(
        cur,
        "kosten",
        "kategorie_de",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS kategorie_de TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE kosten ADD COLUMN kategorie_de TEXT NOT NULL DEFAULT ''",
    )
    ensure_column(
        cur,
        "kosten",
        "beschreibung_de",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS beschreibung_de TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE kosten ADD COLUMN beschreibung_de TEXT NOT NULL DEFAULT ''",
    )
    ensure_column(
        cur,
        "kosten",
        "beleg",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS beleg TEXT",
        "ALTER TABLE kosten ADD COLUMN beleg TEXT",
    )
    ensure_column(
        cur,
        "kosten",
        "ohne_beleg",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS ohne_beleg INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE kosten ADD COLUMN ohne_beleg INTEGER NOT NULL DEFAULT 0",
    )
    ensure_column(
        cur,
        "kosten",
        "genehmigt",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS genehmigt INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE kosten ADD COLUMN genehmigt INTEGER NOT NULL DEFAULT 0",
    )
    ensure_column(
        cur,
        "kosten",
        "genehmigt_von",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS genehmigt_von TEXT",
        "ALTER TABLE kosten ADD COLUMN genehmigt_von TEXT",
    )
    ensure_column(
        cur,
        "kosten",
        "created_at",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS created_at TIMESTAMP",
        "ALTER TABLE kosten ADD COLUMN created_at TEXT",
    )
    ensure_column(
        cur,
        "kosten",
        "reise_name",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS reise_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE kosten ADD COLUMN reise_name TEXT NOT NULL DEFAULT ''",
    )

    # ---- vorschuesse ----
    if DATABASE_URL:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS vorschuesse (
            id BIGSERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            datum DATE,
            reise_name TEXT NOT NULL DEFAULT '',
            beschreibung TEXT NOT NULL DEFAULT '',
            betrag_sar NUMERIC NOT NULL DEFAULT 0,
            beleg TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )
        """)
    else:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS vorschuesse (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            datum TEXT,
            reise_name TEXT NOT NULL DEFAULT '',
            beschreibung TEXT NOT NULL DEFAULT '',
            betrag_sar REAL NOT NULL DEFAULT 0,
            beleg TEXT,
            created_at TEXT
        )
        """)

    # ---- startguthaben migrations for existing databases ----
    ensure_column(
        cur,
        "startguthaben",
        "wechselkurs_beleg",
        "ALTER TABLE startguthaben ADD COLUMN IF NOT EXISTS wechselkurs_beleg TEXT",
        "ALTER TABLE startguthaben ADD COLUMN wechselkurs_beleg TEXT",
    )

    # ---- vorschuesse migrations for existing databases ----
    ensure_column(
        cur,
        "vorschuesse",
        "reise_name",
        "ALTER TABLE vorschuesse ADD COLUMN IF NOT EXISTS reise_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE vorschuesse ADD COLUMN reise_name TEXT NOT NULL DEFAULT ''",
    )
    ensure_column(
        cur,
        "vorschuesse",
        "beschreibung",
        "ALTER TABLE vorschuesse ADD COLUMN IF NOT EXISTS beschreibung TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE vorschuesse ADD COLUMN beschreibung TEXT NOT NULL DEFAULT ''",
    )
    ensure_column(
        cur,
        "vorschuesse",
        "beleg",
        "ALTER TABLE vorschuesse ADD COLUMN IF NOT EXISTS beleg TEXT",
        "ALTER TABLE vorschuesse ADD COLUMN beleg TEXT",
    )
    ensure_column(
        cur,
        "vorschuesse",
        "created_at",
        "ALTER TABLE vorschuesse ADD COLUMN IF NOT EXISTS created_at TIMESTAMP",
        "ALTER TABLE vorschuesse ADD COLUMN created_at TEXT",
    )

    # ---- migration: hide entries from Buchhaltung without deleting for Reiseführer ----
    ensure_column(
        cur,
        "kosten",
        "buchhaltung_deleted",
        "ALTER TABLE kosten ADD COLUMN IF NOT EXISTS buchhaltung_deleted INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE kosten ADD COLUMN buchhaltung_deleted INTEGER NOT NULL DEFAULT 0",
    )

    conn.commit()
    conn.close()


# =====================
# DB schema init (Render-safe)
# =====================
_SCHEMA_READY = False

@app.before_request
def _ensure_schema_once():
    global _SCHEMA_READY
    if _SCHEMA_READY:
        return
    try:
        ensure_schema()
        _SCHEMA_READY = True
    except Exception as e:
        print("DB schema init failed:", repr(e), flush=True)
        raise


# =====================
# SMTP – STRATO
# =====================
def _env_bool(name, default=False):
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in ("1", "true", "yes", "on")


def today_str() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d")


def min_expense_date_str() -> str:
    return (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%d")


def is_expense_date_allowed(date_value: str) -> bool:
    try:
        selected_date = datetime.strptime((date_value or "").strip(), "%Y-%m-%d").date()
    except ValueError:
        return False

    today = datetime.utcnow().date()
    min_allowed = today - timedelta(days=7)
    return min_allowed <= selected_date <= today


def build_reset_link(token: str) -> str:
    base_url = (os.environ.get("RESET_BASE_URL") or os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    if base_url:
        return f"{base_url}{url_for('reset', token=token)}"
    return url_for("reset", token=token, _external=True)


# =====================
# SMTP – allgemein
# =====================
def send_smtp_message(to_email, subject, body):
    host = (os.environ.get("SMTP_HOST") or "").strip()
    port = int(os.environ.get("SMTP_PORT", 587))
    user = (os.environ.get("SMTP_USER") or "").strip()
    password = os.environ.get("SMTP_PASSWORD") or ""
    from_email = (os.environ.get("SMTP_FROM") or user).strip()

    missing = []
    if not host:
        missing.append("SMTP_HOST")
    if not user:
        missing.append("SMTP_USER")
    if not password:
        missing.append("SMTP_PASSWORD")
    if not from_email:
        missing.append("SMTP_FROM")

    if missing:
        raise RuntimeError(
            "SMTP-Konfiguration unvollständig. Fehlend: " + ", ".join(missing)
        )

    msg = MIMEMultipart()
    msg["From"] = from_email
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    use_ssl = _env_bool("SMTP_USE_SSL", port == 465)
    use_tls = _env_bool("SMTP_USE_TLS", not use_ssl)

    print(f"SMTP: connecting to {host}:{port} ssl={use_ssl} tls={use_tls}")
    server = None
    try:
        if use_ssl:
            server = smtplib.SMTP_SSL(host, port, timeout=15)
            server.ehlo()
        else:
            server = smtplib.SMTP(host, port, timeout=15)
            server.ehlo()
            if use_tls:
                server.starttls()
                server.ehlo()

        server.login(user, password)
        server.sendmail(from_email, [to_email], msg.as_string())
        print("SMTP: mail sent")
    finally:
        if server is not None:
            try:
                server.quit()
            except Exception:
                pass


def send_reset_email(to_email, reset_link):
    body = f"""
Hallo,

du kannst dein Passwort über folgenden Link zurücksetzen:

{reset_link}

Der Link ist 30 Minuten gültig.

Falls du das nicht warst, ignoriere diese E-Mail.
"""
    send_smtp_message(to_email, "Passwort zurücksetzen", body)


def send_registration_request_email(pending_user_email, pending_username, requested_role):
    role_label = "Buchhaltung" if requested_role == "buchhaltung" else "Reiseführer"
    approval_link = (os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/") or request.host_url.rstrip("/")
    body = f"""
Hallo,

es gibt eine neue Registrierungsanfrage für die Reisekosten-App.

Benutzername: {pending_username}
E-Mail: {pending_user_email}
Gewünschte Rolle: {role_label}

Bitte melde dich als Buchhalter an und prüfe die Anfrage im Bereich "Registrierungsanfragen":
{approval_link}{url_for('dashboard')}
"""
    send_smtp_message(APPROVER_EMAIL, "Neue Registrierungsanfrage", body)


def send_registration_approved_email(to_email, username, role):
    role_label = "Buchhaltung" if role == "buchhaltung" else "Reiseführer"
    login_link = (os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/") or request.host_url.rstrip("/")
    body = f"""
Hallo {username},

deine Registrierungsanfrage für die Reisekosten-App wurde freigegeben.

Freigegebene Rolle: {role_label}
Login: {login_link}{url_for('login')}

Du kannst dich ab sofort anmelden.
"""
    send_smtp_message(to_email, "Registrierung freigegeben", body)



# =====================
# Routes
# =====================
@app.route("/")
def index():
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    selected_role = request.args.get("role", "reisefuehrer")

    if request.method == "POST":
        email = normalize_email(request.form.get("email"))
        password = request.form.get("password")

        if not email or not password:
            flash("Bitte E-Mail und Passwort eingeben")
            return redirect(url_for("login", role=selected_role))

        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "SELECT id, username, email, password_hash, role, COALESCE(approval_status, 'approved') AS approval_status FROM users WHERE LOWER(email)=LOWER(%s)"
            if DATABASE_URL else
            "SELECT id, username, email, password_hash, role, COALESCE(approval_status, 'approved') AS approval_status FROM users WHERE LOWER(email)=LOWER(?)",
            (email,)
        )
        user = cur.fetchone()
        conn.close()

        if not user:
            flash("Login fehlgeschlagen")
            return redirect(url_for("login", role=selected_role))

        approval_status = row_value(user, "approval_status", user[5] if not hasattr(user, "keys") and len(user) > 5 else "approved")
        if approval_status == "pending":
            flash("Dein Konto wartet noch auf Freigabe.", "err")
            return redirect(url_for("login", role=selected_role))
        if approval_status == "rejected":
            flash("Deine Registrierungsanfrage wurde abgelehnt. Bitte kontaktiere die Buchhaltung.", "err")
            return redirect(url_for("login", role=selected_role))

        pw_hash = user["password_hash"] if hasattr(user, "keys") else user[3]
        if pw_hash and check_password_hash(pw_hash, password):
            session["user_id"] = user["id"] if hasattr(user, "keys") else user[0]
            session["role"] = user["role"] if hasattr(user, "keys") else user[4]
            session["username"] = user["username"] if hasattr(user, "keys") else user[1]
            session["email"] = normalize_email(row_value(user, "email", user[2] if not hasattr(user, "keys") else ""))
            return redirect(url_for("dashboard"))

        flash("Login fehlgeschlagen")
        return redirect(url_for("login", role=selected_role))

    return render_template("login.html", role=selected_role)

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        email = normalize_email(request.form.get("email"))
        password = request.form.get("password") or ""
        password2 = request.form.get("password2") or ""
        role = (request.form.get("role") or "reisefuehrer").strip()
        if role not in ("reisefuehrer", "buchhaltung"):
            role = "reisefuehrer"

        if not username or not email or not password:
            flash("Bitte alle Felder ausfüllen", "err")
            return redirect(url_for("register"))
        if password != password2:
            flash("Die Passwörter stimmen nicht überein.", "err")
            return redirect(url_for("register"))

        pw_hash = generate_password_hash(password)

        conn = get_db()
        cur = conn.cursor()
        try:
            cur.execute(
                "SELECT id, username, email, password_hash, role, COALESCE(approval_status, 'approved') AS approval_status FROM users WHERE LOWER(email) = LOWER(%s)"
                if DATABASE_URL else
                "SELECT id, username, email, password_hash, role, COALESCE(approval_status, 'approved') AS approval_status FROM users WHERE LOWER(email) = LOWER(?)",
                (email,)
            )
            existing_email = cur.fetchone()

            cur.execute(
                "SELECT id, email FROM users WHERE LOWER(username) = LOWER(%s)"
                if DATABASE_URL else
                "SELECT id, email FROM users WHERE LOWER(username) = LOWER(?)",
                (username,)
            )
            existing_username = cur.fetchone()

            if existing_email:
                status = row_value(existing_email, "approval_status", "approved")
                if status == "pending":
                    flash("Für diese E-Mail gibt es bereits eine offene Registrierungsanfrage.", "err")
                    return redirect(url_for("login"))
                if status == "rejected":
                    flash("Diese E-Mail wurde bereits abgelehnt. Bitte kontaktiere die Buchhaltung.", "err")
                    return redirect(url_for("login"))
                flash("E-Mail ist bereits registriert – bitte Passwort zurücksetzen, falls du es vergessen hast.", "err")
                return redirect(url_for("forgot"))

            if existing_username:
                existing_username_email = normalize_email(row_value(existing_username, "email", ""))
                if existing_username_email != email:
                    flash("Der Benutzername ist bereits vergeben.", "err")
                    return redirect(url_for("register"))

            if DATABASE_URL:
                cur.execute(
                    "INSERT INTO users (username, email, password_hash, role, approval_status, approved_at, approved_by_email) VALUES (%s,%s,%s,%s,%s,NULL,NULL)",
                    (username, email, pw_hash, role, "pending")
                )
            else:
                cur.execute(
                    "INSERT INTO users (username, email, password_hash, role, approval_status, approved_at, approved_by_email) VALUES (?,?,?,?,?,NULL,NULL)",
                    (username, email, pw_hash, role, "pending")
                )
            conn.commit()
        finally:
            conn.close()

        flash("Registrierungsanfrage gespeichert. Du kannst dich nach Freigabe anmelden.", "ok")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/forgot", methods=["GET", "POST"])
def forgot():
    if request.method == "POST":
        email = normalize_email(request.form.get("email"))
        if not email:
            flash("Bitte eine E-Mail-Adresse eingeben.", "error")
            return redirect(url_for("forgot"))

        conn = get_db()
        try:
            cur = conn.cursor()
            cur.execute(
                "SELECT * FROM users WHERE LOWER(email) = LOWER(%s)" if DATABASE_URL else
                "SELECT * FROM users WHERE LOWER(email) = LOWER(?)", (email,)
            )
            user = cur.fetchone()

            if user:
                token = uuid.uuid4().hex
                expires = datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(minutes=30)
                expires_db_value = expires if DATABASE_URL else expires.strftime("%Y-%m-%d %H:%M:%S")

                cur.execute(
                    "UPDATE users SET reset_token=%s, reset_token_expires=%s WHERE id=%s"
                    if DATABASE_URL else
                    "UPDATE users SET reset_token=?, reset_token_expires=? WHERE id=?",
                    (token, expires_db_value, row_value(user, "id"))
                )
                conn.commit()

                reset_link = build_reset_link(token)
                try:
                    send_reset_email(row_value(user, "email"), reset_link)
                except Exception as exc:
                    app.logger.exception("Password reset e-mail could not be sent for %s", email)
                    flash(f"Die Reset-E-Mail konnte nicht versendet werden: {exc}", "error")
                    return redirect(url_for("forgot"))
        finally:
            conn.close()

        flash("Wenn die E-Mail existiert, wurde eine Nachricht versendet.")
        return redirect(url_for("login"))

    return render_template("forgot.html")


@app.route("/reset/<token>", methods=["GET", "POST"])
def reset(token):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT * FROM users WHERE reset_token=%s"
            if DATABASE_URL else
            "SELECT * FROM users WHERE reset_token=?",
            (token,)
        )
        user = cur.fetchone()

        expires_at = parse_reset_expiry(row_value(user, "reset_token_expires"))
        if not user or expires_at is None or expires_at < datetime.now(timezone.utc).replace(tzinfo=None):
            flash("Der Reset-Link ist abgelaufen.", "error")
            return redirect(url_for("login"))

        if request.method == "POST":
            password = (request.form.get("password") or "").strip()
            password2 = (request.form.get("password2") or "").strip()
            if not password:
                flash("Bitte ein neues Passwort eingeben.", "error")
                return redirect(url_for("reset", token=token))
            if password != password2:
                flash("Die Passwörter stimmen nicht überein.", "error")
                return redirect(url_for("reset", token=token))

            pw_hash = generate_password_hash(password)
            cur.execute(
                "UPDATE users SET password_hash=%s, reset_token=NULL, reset_token_expires=NULL WHERE id=%s"
                if DATABASE_URL else
                "UPDATE users SET password_hash=?, reset_token=NULL, reset_token_expires=NULL WHERE id=?",
                (pw_hash, row_value(user, "id"))
            )
            conn.commit()

            flash("Passwort erfolgreich geändert")
            return redirect(url_for("login"))

        return render_template("reset.html")
    finally:
        conn.close()


@app.route("/uploads/<path:filename>")
def uploads(filename):
    return send_from_directory(
        app.config["UPLOAD_FOLDER"],
        filename,
        as_attachment=False
    )


def _parse_decimal(val: str) -> float:
    if val is None:
        return 0.0
    val = val.strip().replace(",", ".")
    try:
        return float(val)
    except ValueError:
        return 0.0

def _now_iso():
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def normalize_trip_name(value: str | None) -> str:
    return (value or "").strip()


def month_from_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    text = str(value).strip()
    return text[:7] if len(text) >= 7 else ""


def entry_matches_filters(entry, month_filter="", trip_filter=""):
    if month_filter and month_from_value(entry.get("datum")) != month_filter:
        return False
    if trip_filter and normalize_trip_name(entry.get("reise_name")) != trip_filter:
        return False
    return True


def _status_text(status: int) -> str:
    if status == 1:
        return "Genehmigt"
    if status == -1:
        return "Abgelehnt"
    return "Offen"


def fetch_startbetrag(cur, user_id):
    cur.execute(
        "SELECT betrag_sar FROM startguthaben WHERE user_id=%s" if DATABASE_URL else
        "SELECT betrag_sar FROM startguthaben WHERE user_id=?",
        (user_id,),
    )
    row = cur.fetchone()
    return float(row_value(row, "betrag_sar", 0) if row else 0)


def fetch_wechselkurs_beleg(cur, user_id):
    cur.execute(
        "SELECT wechselkurs_beleg FROM startguthaben WHERE user_id=%s" if DATABASE_URL else
        "SELECT wechselkurs_beleg FROM startguthaben WHERE user_id=?",
        (user_id,),
    )
    row = cur.fetchone()
    return row_value(row, "wechselkurs_beleg") if row else None


def is_missing_db_object_error(exc, *names):
    message = str(exc).lower()
    return any(name.lower() in message for name in names)


def fetch_users(cur):
    cur.execute(
        "SELECT id, username FROM users WHERE role=%s AND COALESCE(approval_status, 'approved')=%s ORDER BY username" if DATABASE_URL else
        "SELECT id, username FROM users WHERE role=? AND COALESCE(approval_status, 'approved')=? ORDER BY username",
        ("reisefuehrer", "approved"),
    )
    rows = cur.fetchall() or []
    result = []
    for row in rows:
        result.append({
            "id": row_value(row, "id", row[0] if not hasattr(row, "keys") else None),
            "username": row_value(row, "username", row[1] if not hasattr(row, "keys") else ""),
        })
    return result

def is_request_approver():
    return session.get("role") == "buchhaltung" and normalize_email(session.get("email")) == APPROVER_EMAIL


def fetch_pending_registrations(cur):
    cur.execute(
        "SELECT id, username, email, role, COALESCE(approval_status, 'approved') AS approval_status FROM users WHERE COALESCE(approval_status, 'approved')=%s ORDER BY id DESC"
        if DATABASE_URL else
        "SELECT id, username, email, role, COALESCE(approval_status, 'approved') AS approval_status FROM users WHERE COALESCE(approval_status, 'approved')=? ORDER BY id DESC",
        ("pending",),
    )
    rows = cur.fetchall() or []
    result = []
    for row in rows:
        result.append({
            "id": row_value(row, "id"),
            "username": row_value(row, "username", ""),
            "email": normalize_email(row_value(row, "email", "")),
            "role": row_value(row, "role", "reisefuehrer"),
            "approval_status": row_value(row, "approval_status", "pending"),
        })
    return result


def fetch_kosten(cur, user_id=None, include_username=False):
    rows_raw = None
    try:
        if include_username:
            query = """
                SELECT k.id, k.user_id, u.username, k.datum, k.kategorie_ar, k.kategorie_de,
                       k.beschreibung_ar, k.beschreibung_de, k.betrag_sar, k.beleg,
                       k.genehmigt, k.genehmigt_von, k.reise_name
                FROM kosten k
                JOIN users u ON u.id = k.user_id
                WHERE COALESCE(k.buchhaltung_deleted,0)=0
            """
            params = []
            if user_id is not None:
                query += " AND k.user_id=%s" if DATABASE_URL else " AND k.user_id=?"
                params.append(user_id)
            query += " ORDER BY k.datum DESC, k.id DESC"
            cur.execute(query, tuple(params))
        else:
            cur.execute(
                "SELECT id, user_id, datum, kategorie_ar, kategorie_de, beschreibung_ar, beschreibung_de, betrag_sar, beleg, genehmigt, genehmigt_von, reise_name FROM kosten WHERE user_id=%s ORDER BY datum DESC, id DESC"
                if DATABASE_URL else
                "SELECT id, user_id, datum, kategorie_ar, kategorie_de, beschreibung_ar, beschreibung_de, betrag_sar, beleg, genehmigt, genehmigt_von, reise_name FROM kosten WHERE user_id=? ORDER BY datum DESC, id DESC",
                (user_id,),
            )
        rows_raw = cur.fetchall() or []
    except Exception as exc:
        if not is_missing_db_object_error(exc, 'reise_name', 'buchhaltung_deleted'):
            raise
        # Fallback for old databases where the migration has not run yet.
        if include_username:
            query = """
                SELECT k.id, k.user_id, u.username, k.datum, k.kategorie_ar, k.kategorie_de,
                       k.beschreibung_ar, k.beschreibung_de, k.betrag_sar, k.beleg,
                       k.genehmigt, k.genehmigt_von
                FROM kosten k
                JOIN users u ON u.id = k.user_id
                ORDER BY k.datum DESC, k.id DESC
            """
            params = []
            if user_id is not None:
                query = query.replace('ORDER BY', ("WHERE k.user_id=%s ORDER BY" if DATABASE_URL else "WHERE k.user_id=? ORDER BY"), 1)
                params.append(user_id)
            cur.execute(query, tuple(params))
        else:
            cur.execute(
                "SELECT id, user_id, datum, kategorie_ar, kategorie_de, beschreibung_ar, beschreibung_de, betrag_sar, beleg, genehmigt, genehmigt_von FROM kosten WHERE user_id=%s ORDER BY datum DESC, id DESC"
                if DATABASE_URL else
                "SELECT id, user_id, datum, kategorie_ar, kategorie_de, beschreibung_ar, beschreibung_de, betrag_sar, beleg, genehmigt, genehmigt_von FROM kosten WHERE user_id=? ORDER BY datum DESC, id DESC",
                (user_id,),
            )
        rows_raw = cur.fetchall() or []

    rows = []
    for row in rows_raw:
        rows.append({
            "id": row_value(row, "id", row[0] if not hasattr(row, "keys") else None),
            "user_id": row_value(row, "user_id", row[1] if not hasattr(row, "keys") else None),
            "username": row_value(row, "username", ""),
            "datum": str(row_value(row, "datum", row[2] if not hasattr(row, "keys") else ""))[:10],
            "kategorie_ar": resolve_category(row_value(row, "kategorie_ar", ""))[0],
            "kategorie_de": row_value(row, "kategorie_de", "") or resolve_category(row_value(row, "kategorie_ar", ""))[1],
            "beschreibung_ar": row_value(row, "beschreibung_ar", ""),
            "beschreibung_de": row_value(row, "beschreibung_de", ""),
            "betrag_sar": float(row_value(row, "betrag_sar", 0) or 0),
            "beleg": row_value(row, "beleg"),
            "genehmigt": int(row_value(row, "genehmigt", 0) or 0),
            "genehmigt_von": row_value(row, "genehmigt_von", ""),
            "reise_name": normalize_trip_name(row_value(row, "reise_name", "")),
        })
    return rows


def fetch_vorschuesse(cur, user_id=None, include_username=False):
    if include_username:
        query = """
            SELECT v.id, v.user_id, u.username, v.datum, v.reise_name, v.beschreibung,
                   v.betrag_sar, v.beleg
            FROM vorschuesse v
            JOIN users u ON u.id = v.user_id
        """
        params = []
        if user_id is not None:
            query += " WHERE v.user_id=%s" if DATABASE_URL else " WHERE v.user_id=?"
            params.append(user_id)
        query += " ORDER BY v.datum DESC, v.id DESC"
        cur.execute(query, tuple(params))
    else:
        cur.execute(
            "SELECT id, user_id, datum, reise_name, beschreibung, betrag_sar, beleg FROM vorschuesse WHERE user_id=%s ORDER BY datum DESC, id DESC"
            if DATABASE_URL else
            "SELECT id, user_id, datum, reise_name, beschreibung, betrag_sar, beleg FROM vorschuesse WHERE user_id=? ORDER BY datum DESC, id DESC",
            (user_id,),
        )
    rows = []
    for row in (cur.fetchall() or []):
        rows.append({
            "id": row_value(row, "id", row[0] if not hasattr(row, "keys") else None),
            "user_id": row_value(row, "user_id", row[1] if not hasattr(row, "keys") else None),
            "username": row_value(row, "username", ""),
            "datum": str(row_value(row, "datum", row[2] if not hasattr(row, "keys") else ""))[:10],
            "reise_name": normalize_trip_name(row_value(row, "reise_name", row[3] if not hasattr(row, "keys") else "")),
            "beschreibung": row_value(row, "beschreibung", row[4] if not hasattr(row, "keys") else ""),
            "betrag_sar": float(row_value(row, "betrag_sar", row[5] if not hasattr(row, "keys") else 0) or 0),
            "beleg": row_value(row, "beleg", row[6] if not hasattr(row, "keys") else None),
        })
    return rows


def unique_trip_names(*collections):
    values = set()
    for collection in collections:
        for row in collection:
            name = normalize_trip_name(row.get("reise_name"))
            if name:
                values.add(name)
    return sorted(values, key=lambda x: x.lower())


def compute_scope_summary(start_amount, kosten_rows, vorschuss_rows, month_filter="", trip_filter=""):
    filtered_kosten = [r for r in kosten_rows if r.get("genehmigt") != -1 and entry_matches_filters(r, month_filter, trip_filter)]
    filtered_vorschuesse = [r for r in vorschuss_rows if entry_matches_filters(r, month_filter, trip_filter)]

    if month_filter:
        previous_kosten = [
            r for r in kosten_rows
            if r.get("genehmigt") != -1
            and (not trip_filter or normalize_trip_name(r.get("reise_name")) == trip_filter)
            and month_from_value(r.get("datum")) < month_filter
        ]
        previous_vorschuesse = [
            r for r in vorschuss_rows
            if (not trip_filter or normalize_trip_name(r.get("reise_name")) == trip_filter)
            and month_from_value(r.get("datum")) < month_filter
        ]
        opening = (0.0 if trip_filter else start_amount) + sum(r["betrag_sar"] for r in previous_vorschuesse) - sum(r["betrag_sar"] for r in previous_kosten)
    else:
        opening = 0.0 if trip_filter else start_amount

    vorschuss_total = sum(r["betrag_sar"] for r in filtered_vorschuesse)
    kosten_total = sum(r["betrag_sar"] for r in filtered_kosten)
    closing = opening + vorschuss_total - kosten_total

    return {
        "opening": opening,
        "vorschuss_total": vorschuss_total,
        "kosten_total": kosten_total,
        "closing": closing,
        "filtered_kosten": filtered_kosten,
        "filtered_vorschuesse": filtered_vorschuesse,
    }


def aggregate_summaries(user_ids, start_map, kosten_rows, vorschuss_rows, month_filter="", trip_filter=""):
    summary = {"opening": 0.0, "vorschuss_total": 0.0, "kosten_total": 0.0, "closing": 0.0}
    for user_id in user_ids:
        user_kosten = [row for row in kosten_rows if row["user_id"] == user_id]
        user_vorschuesse = [row for row in vorschuss_rows if row["user_id"] == user_id]
        user_summary = compute_scope_summary(start_map.get(user_id, 0.0), user_kosten, user_vorschuesse, month_filter, trip_filter)
        for key in summary:
            summary[key] += user_summary[key]
    return summary


@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if "user_id" not in session:
        return redirect(url_for("login"))

    role = session.get("role", "reisefuehrer")
    uid = session["user_id"]

    if request.method == "POST":
        action = request.form.get("action")

        if role == "buchhaltung":
            kid = request.form.get("kid")
            pending_user_id = request.form.get("pending_user_id")
            conn = get_db()
            cur = conn.cursor()
            try:
                if action == "translate_descriptions":
                    month_filter_post = (request.args.get("monat") or "").strip()
                    trip_filter_post = normalize_trip_name(request.args.get("reise"))
                    selected_user_id_raw_post = (request.args.get("user_id") or "").strip()
                    selected_user_id_post = int(selected_user_id_raw_post) if selected_user_id_raw_post.isdigit() else None
                    rows_to_translate = fetch_kosten(cur, include_username=True)
                    if selected_user_id_post:
                        rows_to_translate = [row for row in rows_to_translate if row["user_id"] == selected_user_id_post]
                    rows_to_translate = [row for row in rows_to_translate if entry_matches_filters(row, month_filter_post, trip_filter_post)]
                    translated_count = translate_and_store_missing_descriptions(
                        cur, rows_to_translate, max_items=8, time_budget_seconds=12
                    )
                    conn.commit()
                    if translated_count:
                        flash(f"{translated_count} Beschreibung(en) wurden übersetzt. Wenn noch arabische Texte vorhanden sind, den Button erneut klicken.", "ok")
                    else:
                        flash("Keine neuen Übersetzungen gefunden oder der Übersetzungsdienst hat nicht schnell genug geantwortet.", "err")
                elif action in ("approve", "deny") and kid:
                    genehmigt_val = 1 if action == "approve" else -1
                    cur.execute(
                        "UPDATE kosten SET genehmigt=%s, genehmigt_von=%s WHERE id=%s" if DATABASE_URL else
                        "UPDATE kosten SET genehmigt=?, genehmigt_von=? WHERE id=?",
                        (genehmigt_val, session.get("username", ""), kid),
                    )
                    conn.commit()
                    flash("Status aktualisiert", "ok")
                elif action == "delete_admin" and kid:
                    cur.execute(
                        "DELETE FROM kosten WHERE id=%s" if DATABASE_URL else
                        "DELETE FROM kosten WHERE id=?",
                        (kid,),
                    )
                    conn.commit()
                    flash("Eintrag gelöscht", "ok")
                elif action in ("approve_registration", "reject_registration") and pending_user_id:
                    if not is_request_approver():
                        flash("Nur der freigegebene Buchhalter darf Registrierungen bearbeiten.", "err")
                    else:
                        cur.execute(
                            "SELECT id, username, email, role, COALESCE(approval_status, 'approved') AS approval_status FROM users WHERE id=%s" if DATABASE_URL else
                            "SELECT id, username, email, role, COALESCE(approval_status, 'approved') AS approval_status FROM users WHERE id=?",
                            (pending_user_id,),
                        )
                        pending_user = cur.fetchone()
                        if not pending_user:
                            flash("Anfrage nicht gefunden.", "err")
                        elif row_value(pending_user, "approval_status", "approved") != "pending":
                            flash("Diese Anfrage wurde bereits bearbeitet.", "err")
                        else:
                            new_status = "approved" if action == "approve_registration" else "rejected"
                            approved_at_value = datetime.utcnow() if DATABASE_URL else _now_iso()
                            cur.execute(
                                "UPDATE users SET approval_status=%s, approved_at=%s, approved_by_email=%s WHERE id=%s" if DATABASE_URL else
                                "UPDATE users SET approval_status=?, approved_at=?, approved_by_email=? WHERE id=?",
                                (new_status, approved_at_value, normalize_email(session.get("email")), pending_user_id),
                            )
                            conn.commit()
                            if new_status == "approved":
                                try:
                                    send_registration_approved_email(
                                        normalize_email(row_value(pending_user, "email", "")),
                                        row_value(pending_user, "username", ""),
                                        row_value(pending_user, "role", "reisefuehrer"),
                                    )
                                    flash("Registrierung freigegeben und E-Mail versendet.", "ok")
                                except Exception as exc:
                                    app.logger.exception("Approval confirmation e-mail could not be sent for user %s", pending_user_id)
                                    flash(f"Registrierung freigegeben, aber die Bestätigungs-E-Mail konnte nicht versendet werden: {exc}", "err")
                            else:
                                flash("Registrierungsanfrage wurde abgelehnt und bleibt blockiert.", "ok")
            finally:
                conn.close()
            return redirect(url_for("dashboard", user_id=request.args.get("user_id", ""), monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

        conn = get_db()
        cur = conn.cursor()
        try:
            if action == "set_start":
                betrag = _parse_decimal(request.form.get("startguthaben_sar", "0"))
                wechselkurs_beleg = None
                file = get_first_uploaded_file("wechselkurs_beleg_camera", "wechselkurs_beleg_library", "wechselkurs_beleg")
                try:
                    wechselkurs_beleg = save_uploaded_file(file, "wechselkurs")
                except RuntimeError as exc:
                    flash(str(exc), "err")
                    return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

                if DATABASE_URL:
                    cur.execute(
                        """
                        INSERT INTO startguthaben (user_id, betrag_sar, wechselkurs_beleg)
                        VALUES (%s,%s,%s)
                        ON CONFLICT (user_id) DO UPDATE SET
                            betrag_sar=EXCLUDED.betrag_sar,
                            wechselkurs_beleg=COALESCE(EXCLUDED.wechselkurs_beleg, startguthaben.wechselkurs_beleg)
                        """,
                        (uid, betrag, wechselkurs_beleg),
                    )
                else:
                    cur.execute("SELECT wechselkurs_beleg FROM startguthaben WHERE user_id=?", (uid,))
                    existing = cur.fetchone()
                    existing_beleg = row_value(existing, "wechselkurs_beleg", existing[0] if existing else None)
                    cur.execute(
                        "INSERT OR REPLACE INTO startguthaben (user_id, betrag_sar, wechselkurs_beleg) VALUES (?,?,?)",
                        (uid, betrag, wechselkurs_beleg or existing_beleg),
                    )
                conn.commit()
                flash("Anfangsbestand gespeichert", "ok")
                return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

            if action == "reset_start":
                cur.execute(
                    "DELETE FROM startguthaben WHERE user_id=%s" if DATABASE_URL else
                    "DELETE FROM startguthaben WHERE user_id=?",
                    (uid,),
                )
                conn.commit()
                flash("Anfangsbestand zurückgesetzt", "ok")
                return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

            if action == "add_vorschuss":
                datum = request.form.get("vorschuss_datum") or datetime.utcnow().strftime("%Y-%m-%d")
                reise_name = normalize_trip_name(request.form.get("vorschuss_reise"))
                beschreibung = (request.form.get("vorschuss_beschreibung") or "").strip()
                betrag = _parse_decimal(request.form.get("vorschuss_betrag_sar", "0"))
                if betrag <= 0:
                    flash("Bitte einen gültigen Vorschussbetrag eingeben.", "err")
                    return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

                beleg_filename = None
                beleg_file = get_first_uploaded_file("vorschuss_beleg_camera", "vorschuss_beleg_library", "vorschuss_beleg")
                try:
                    beleg_filename = save_uploaded_file(beleg_file, "vorschuss")
                except RuntimeError as exc:
                    flash(str(exc), "err")
                    return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

                if DATABASE_URL:
                    cur.execute(
                        "INSERT INTO vorschuesse (user_id, datum, reise_name, beschreibung, betrag_sar, beleg, created_at) VALUES (%s,%s,%s,%s,%s,%s,NOW())",
                        (uid, datum, reise_name, beschreibung, betrag, beleg_filename),
                    )
                else:
                    cur.execute(
                        "INSERT INTO vorschuesse (user_id, datum, reise_name, beschreibung, betrag_sar, beleg, created_at) VALUES (?,?,?,?,?,?,?)",
                        (uid, datum, reise_name, beschreibung, betrag, beleg_filename, _now_iso()),
                    )
                conn.commit()
                flash("Vorschuss gespeichert", "ok")
                return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

            if action == "delete_vorschuss":
                vid = request.form.get("vid")
                if vid:
                    cur.execute(
                        "DELETE FROM vorschuesse WHERE id=%s AND user_id=%s" if DATABASE_URL else
                        "DELETE FROM vorschuesse WHERE id=? AND user_id=?",
                        (vid, uid),
                    )
                    conn.commit()
                    flash("Vorschuss gelöscht", "ok")
                return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

            if action == "add_kosten":
                datum = request.form.get("datum")
                if not is_expense_date_allowed(datum):
                    flash("Ausgaben können nur für heute oder maximal 7 Tage rückwirkend eingetragen werden.", "err")
                    return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

                kategorie_raw = request.form.get("kategorie")
                kategorie_ar, kategorie_de = resolve_category(kategorie_raw)
                beschreibung_ar = (request.form.get("beschreibung_ar") or "").strip()
                beschreibung_de = "" if contains_arabic(beschreibung_ar) else beschreibung_ar
                betrag = _parse_decimal(request.form.get("betrag_sar", "0"))
                reise_name = normalize_trip_name(request.form.get("reise_name"))

                # Wichtig für Render/PostgreSQL:
                # In vorhandenen Render-Datenbanken ist kosten.ohne_beleg teilweise BOOLEAN.
                # PostgreSQL akzeptiert dafür kein 0/1. Lokal mit SQLite bleibt 0/1 korrekt.
                ohne_beleg_checked = bool(request.form.get("ohne_beleg"))
                ohne_beleg = ohne_beleg_checked if DATABASE_URL else (1 if ohne_beleg_checked else 0)

                if betrag <= 0:
                    flash("Bitte einen gültigen Ausgabenbetrag eingeben.", "err")
                    return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

                beleg_filename = None
                beleg_file = get_first_uploaded_file("beleg_camera", "beleg_library", "beleg")
                try:
                    beleg_filename = save_uploaded_file(beleg_file, "beleg")
                except RuntimeError as exc:
                    flash(str(exc), "err")
                    return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

                if not ohne_beleg_checked and not beleg_filename:
                    flash("Bitte Beleg hochladen oder 'بدون إيصال' aktivieren.", "err")
                    return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

                if DATABASE_URL:
                    cur.execute(
                        """
                        INSERT INTO kosten (
                            user_id, datum, kategorie_ar, kategorie_de, beschreibung_ar, beschreibung_de,
                            betrag_sar, beleg, ohne_beleg, genehmigt, created_at, reise_name
                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,0,NOW(),%s)
                        """,
                        (uid, datum, kategorie_ar, kategorie_de, beschreibung_ar, beschreibung_de, betrag, beleg_filename, ohne_beleg, reise_name),
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO kosten (
                            user_id, datum, kategorie_ar, kategorie_de, beschreibung_ar, beschreibung_de,
                            betrag_sar, beleg, ohne_beleg, genehmigt, created_at, reise_name
                        )
                        VALUES (?,?,?,?,?,?,?,?,?,0,?,?)
                        """,
                        (uid, datum, kategorie_ar, kategorie_de, beschreibung_ar, beschreibung_de, betrag, beleg_filename, ohne_beleg, _now_iso(), reise_name),
                    )
                conn.commit()
                flash("Ausgabe gespeichert", "ok")
                return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

            if action == "delete_kosten":
                kid = request.form.get("kid")
                if kid:
                    cur.execute(
                        "DELETE FROM kosten WHERE id=%s AND user_id=%s AND genehmigt=0" if DATABASE_URL else
                        "DELETE FROM kosten WHERE id=? AND user_id=? AND genehmigt=0",
                        (kid, uid),
                    )
                    conn.commit()
                    flash("Gelöscht", "ok")
                return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))
        except Exception as exc:
            # Wichtig für Render: konkrete Fehler beim Speichern sichtbar loggen und nicht als generischen 500-Fehler anzeigen.
            try:
                conn.rollback()
            except Exception:
                pass
            app.logger.exception(
                "Fehler im Reisefuehrer-Dashboard POST action=%s user_id=%s form_keys=%s files=%s",
                action,
                uid,
                sorted(request.form.keys()),
                sorted(request.files.keys()),
            )
            flash(f"Beim Speichern ist ein Fehler aufgetreten: {exc}", "err")
            return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))
        finally:
            conn.close()

        return redirect(url_for("dashboard", monat=request.args.get("monat", ""), reise=request.args.get("reise", "")))

    month_filter = (request.args.get("monat") or "").strip()
    trip_filter = normalize_trip_name(request.args.get("reise"))

    conn = get_db()
    cur = conn.cursor()
    try:
        if role == "buchhaltung":
            users = fetch_users(cur)
            selected_user_id_raw = (request.args.get("user_id") or "").strip()
            selected_user_id = int(selected_user_id_raw) if selected_user_id_raw.isdigit() else None

            kosten_all = fetch_kosten(cur, include_username=True)
            vorschuesse_all = fetch_vorschuesse(cur, include_username=True)
            if selected_user_id:
                kosten_all = [row for row in kosten_all if row["user_id"] == selected_user_id]
                vorschuesse_all = [row for row in vorschuesse_all if row["user_id"] == selected_user_id]

            trip_options = unique_trip_names(kosten_all, vorschuesse_all)
            filtered_kosten = [row for row in kosten_all if entry_matches_filters(row, month_filter, trip_filter)]
            filtered_vorschuesse = [row for row in vorschuesse_all if entry_matches_filters(row, month_filter, trip_filter)]
            missing_translation_count = sum(
                1 for row in filtered_kosten
                if translation_is_missing_or_not_german(row.get("beschreibung_ar"), row.get("beschreibung_de"))
            )

            start_map = {user["id"]: fetch_startbetrag(cur, user["id"]) for user in users}
            users_summary = []
            for user in users:
                user_start = start_map.get(user["id"], 0.0)
                user_kosten = fetch_kosten(cur, user_id=user["id"], include_username=True)
                user_vorschuesse = fetch_vorschuesse(cur, user_id=user["id"], include_username=True)
                adv_total = sum(row["betrag_sar"] for row in user_vorschuesse)
                exp_total = sum(row["betrag_sar"] for row in user_kosten if row.get("genehmigt") != -1)
                users_summary.append({
                    "id": user["id"],
                    "username": user["username"],
                    "start": user_start,
                    "vorschuesse": adv_total,
                    "total": exp_total,
                    "saldo": user_start + adv_total - exp_total,
                })

            visible_user_ids = [selected_user_id] if selected_user_id else [user["id"] for user in users]
            scope_summary = aggregate_summaries(visible_user_ids, start_map, kosten_all, vorschuesse_all, month_filter, trip_filter)
            can_manage_requests = is_request_approver()
            pending_registrations = fetch_pending_registrations(cur) if can_manage_requests else []

            return render_template(
                "admin.html",
                users_summary=users_summary,
                kosten=filtered_kosten,
                vorschuesse=filtered_vorschuesse,
                trip_options=trip_options,
                selected_month=month_filter,
                selected_trip=trip_filter,
                selected_user_id=selected_user_id_raw,
                users=users,
                scope_summary=scope_summary,
                can_manage_requests=can_manage_requests,
                pending_registrations=pending_registrations,
                approver_email=APPROVER_EMAIL,
                missing_translation_count=missing_translation_count,
            )

        start_amount = fetch_startbetrag(cur, uid)
        wechselkurs_beleg = fetch_wechselkurs_beleg(cur, uid)
        kosten_rows = fetch_kosten(cur, user_id=uid)
        vorschuesse_rows = fetch_vorschuesse(cur, user_id=uid)
        trip_options = unique_trip_names(kosten_rows, vorschuesse_rows)
        scope_summary = compute_scope_summary(start_amount, kosten_rows, vorschuesse_rows, month_filter, trip_filter)

        return render_template(
            "dashboard.html",
            ordered=list(CATEGORIES.items()),
            today=today_str(),
            min_expense_date=min_expense_date_str(),
            start=start_amount,
            opening=scope_summary["opening"],
            vorschuss_total=scope_summary["vorschuss_total"],
            total=scope_summary["kosten_total"],
            saldo=scope_summary["closing"],
            rows=scope_summary["filtered_kosten"],
            vorschuesse=scope_summary["filtered_vorschuesse"],
            wechselkurs_beleg=wechselkurs_beleg,
            trip_options=trip_options,
            selected_month=month_filter,
            selected_trip=trip_filter,
        )
    finally:
        conn.close()


@app.route("/export_excel")
def export_excel():
    if "user_id" not in session:
        return redirect(url_for("login"))

    role = session.get("role", "reisefuehrer")
    month_filter = (request.args.get("monat") or "").strip()
    trip_filter = normalize_trip_name(request.args.get("reise"))
    selected_user_id_raw = (request.args.get("user_id") or "").strip()
    selected_user_id = int(selected_user_id_raw) if selected_user_id_raw.isdigit() else None

    conn = get_db()
    cur = conn.cursor()
    try:
        if role == "buchhaltung":
            users = fetch_users(cur)
            start_map = {user["id"]: fetch_startbetrag(cur, user["id"]) for user in users}
            kosten_rows = fetch_kosten(cur, include_username=True)
            vorschuesse_rows = fetch_vorschuesse(cur, include_username=True)
            if selected_user_id:
                kosten_rows = [row for row in kosten_rows if row["user_id"] == selected_user_id]
                vorschuesse_rows = [row for row in vorschuesse_rows if row["user_id"] == selected_user_id]
                visible_user_ids = [selected_user_id]
                title_user = next((u["username"] for u in users if u["id"] == selected_user_id), f"user-{selected_user_id}")
            else:
                visible_user_ids = [user["id"] for user in users]
                title_user = "alle"
            summary = aggregate_summaries(visible_user_ids, start_map, kosten_rows, vorschuesse_rows, month_filter, trip_filter)
            filtered_kosten = [row for row in kosten_rows if entry_matches_filters(row, month_filter, trip_filter)]
            filtered_vorschuesse = [row for row in vorschuesse_rows if entry_matches_filters(row, month_filter, trip_filter)]
        else:
            title_user = session.get("username", "reisefuehrer")
            start_amount = fetch_startbetrag(cur, session["user_id"])
            kosten_rows = fetch_kosten(cur, user_id=session["user_id"])
            vorschuesse_rows = fetch_vorschuesse(cur, user_id=session["user_id"])
            summary_data = compute_scope_summary(start_amount, kosten_rows, vorschuesse_rows, month_filter, trip_filter)
            summary = {
                "opening": summary_data["opening"],
                "vorschuss_total": summary_data["vorschuss_total"],
                "kosten_total": summary_data["kosten_total"],
                "closing": summary_data["closing"],
            }
            filtered_kosten = summary_data["filtered_kosten"]
            filtered_vorschuesse = summary_data["filtered_vorschuesse"]

        # Echte Excel-Datei erzeugen. Wichtig: Beträge werden als Zahlen geschrieben,
        # nicht als Text. Dadurch funktionieren AutoSumme, Filter und eigene Formeln in Excel.
        wb = Workbook()
        ws = wb.active
        ws.title = "Bericht"

        money_fmt = '#,##0.00'
        title_fill = PatternFill("solid", fgColor="FF1F2937")
        section_fill = PatternFill("solid", fgColor="FFE5E7EB")
        total_fill = PatternFill("solid", fgColor="FFFFF2CC")
        thin_gray = Side(style="thin", color="FFD1D5DB")
        border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

        def style_range(row_idx, start_col, end_col, fill=None, bold=False, font_color="FF111827"):
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if fill:
                    cell.fill = fill
                if bold:
                    cell.font = Font(bold=True, color=font_color)

        def write_header(row_idx, values):
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            style_range(row_idx, 1, len(values), fill=section_fill, bold=True)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        ws["A1"] = "Reisekosten-Bericht"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFFFF")
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center")

        meta_rows = [
            ("Benutzer", title_user),
            ("Monat", month_filter or "alle"),
            ("Reise", trip_filter or "alle"),
            ("Exportiert am", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ]
        for idx, (label, value) in enumerate(meta_rows, start=3):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)

        summary_header_row = 8
        write_header(summary_header_row, ["Kennzahl", "Betrag (SAR)"])
        opening_row = summary_header_row + 1
        advances_summary_row = summary_header_row + 2
        expenses_summary_row = summary_header_row + 3
        closing_row = summary_header_row + 4

        ws.cell(row=opening_row, column=1, value="Anfangssaldo")
        ws.cell(row=opening_row, column=2, value=float(summary.get("opening", 0) or 0))
        ws.cell(row=advances_summary_row, column=1, value="Vorschüsse")
        ws.cell(row=expenses_summary_row, column=1, value="Ausgaben")
        ws.cell(row=closing_row, column=1, value="Restguthaben")
        for row_idx in range(opening_row, closing_row + 1):
            style_range(row_idx, 1, 2, fill=total_fill if row_idx == closing_row else None, bold=(row_idx == closing_row))
            ws.cell(row=row_idx, column=2).number_format = money_fmt

        current_row = closing_row + 3

        # Vorschüsse
        ws.cell(row=current_row, column=1, value="Vorschüsse")
        style_range(current_row, 1, 7, fill=title_fill, bold=True, font_color="FFFFFFFF")
        current_row += 1
        write_header(current_row, ["ID", "Reiseführer", "Datum", "Reise", "Beschreibung", "Betrag (SAR)", "Beleg"])
        vorschuss_data_start = current_row + 1
        current_row = vorschuss_data_start
        if filtered_vorschuesse:
            for row in filtered_vorschuesse:
                values = [
                    row.get("id"),
                    row.get("username") or title_user,
                    row.get("datum"),
                    row.get("reise_name") or "",
                    row.get("beschreibung") or "",
                    float(row.get("betrag_sar", 0) or 0),
                    row.get("beleg") or "",
                ]
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=current_row, column=col_idx, value=value)
                ws.cell(row=current_row, column=6).number_format = money_fmt
                style_range(current_row, 1, 7)
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value="Keine Vorschüsse im gewählten Filter")
            style_range(current_row, 1, 7)
            current_row += 1
        vorschuss_data_end = current_row - 1
        vorschuss_total_row = current_row
        ws.cell(row=vorschuss_total_row, column=5, value="Summe Vorschüsse")
        ws.cell(row=vorschuss_total_row, column=6, value=f"=SUM(F{vorschuss_data_start}:F{vorschuss_data_end})")
        ws.cell(row=vorschuss_total_row, column=6).number_format = money_fmt
        style_range(vorschuss_total_row, 1, 7, fill=total_fill, bold=True)

        current_row += 3

        # Ausgaben
        ws.cell(row=current_row, column=1, value="Ausgaben")
        style_range(current_row, 1, 9, fill=title_fill, bold=True, font_color="FFFFFFFF")
        current_row += 1
        write_header(current_row, ["ID", "Reiseführer", "Datum", "Reise", "Kategorie", "Beschreibung (Deutsch)", "Betrag (SAR)", "Status", "Beleg"])
        kosten_data_start = current_row + 1
        current_row = kosten_data_start
        if filtered_kosten:
            for row in filtered_kosten:
                values = [
                    row.get("id"),
                    row.get("username") or title_user,
                    row.get("datum"),
                    row.get("reise_name") or "",
                    row.get("kategorie_de") or row.get("kategorie_ar") or "",
                    row.get("beschreibung_de") or row.get("beschreibung_ar") or "",
                    float(row.get("betrag_sar", 0) or 0),
                    _status_text(int(row.get("genehmigt", 0) or 0)),
                    row.get("beleg") or "",
                ]
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=current_row, column=col_idx, value=value)
                ws.cell(row=current_row, column=7).number_format = money_fmt
                style_range(current_row, 1, 9)
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value="Keine Ausgaben im gewählten Filter")
            style_range(current_row, 1, 9)
            current_row += 1
        kosten_data_end = current_row - 1
        kosten_total_row = current_row
        ws.cell(row=kosten_total_row, column=6, value="Summe Ausgaben")
        ws.cell(row=kosten_total_row, column=7, value=f"=SUM(G{kosten_data_start}:G{kosten_data_end})")
        ws.cell(row=kosten_total_row, column=7).number_format = money_fmt
        style_range(kosten_total_row, 1, 9, fill=total_fill, bold=True)

        # Zusammenfassung oben per Excel-Formeln anbinden.
        ws.cell(row=advances_summary_row, column=2, value=f"=F{vorschuss_total_row}")
        ws.cell(row=expenses_summary_row, column=2, value=f"=G{kosten_total_row}")
        ws.cell(row=closing_row, column=2, value=f"=B{opening_row}+B{advances_summary_row}-B{expenses_summary_row}")

        for col_idx, width in enumerate([10, 18, 14, 18, 24, 36, 16, 16, 28], start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = f"A{vorschuss_data_start}"

        file_parts = ["reisekosten_bericht", title_user]
        if month_filter:
            file_parts.append(month_filter)
        if trip_filter:
            file_parts.append(trip_filter.replace(" ", "_"))
        filename = secure_filename("_".join(file_parts)) + ".xlsx"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        conn.close()

# =========================
# ✅ NEU: Komplettes Backup (DB + uploads) als ZIP
# =========================
@app.route("/admin/export-backup")
def export_backup():
    if session.get("role") != "buchhaltung":
        flash("Keine Berechtigung.", "error")
        return redirect(url_for("dashboard"))

    def serialize_backup_value(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        return str(value)

    def write_table_csv(zipf, table_name, rows):
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")

        if rows:
            columns = list(rows[0].keys()) if hasattr(rows[0], "keys") else list(range(len(rows[0])))
            writer.writerow(columns)
            for row in rows:
                if hasattr(row, "keys"):
                    writer.writerow([serialize_backup_value(row.get(col)) for col in columns])
                else:
                    writer.writerow([serialize_backup_value(value) for value in row])
        else:
            writer.writerow(["info"])
            writer.writerow([f"Keine Daten in Tabelle {table_name}."])

        zipf.writestr(f"database/{table_name}.csv", buffer.getvalue().encode("utf-8-sig"))

    ts = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
    zip_name = f"reisekosten_backup_{ts}.zip"
    tmp_zip = tempfile.NamedTemporaryFile(prefix="reisekosten_backup_", suffix=".zip", delete=False)
    zip_path = tmp_zip.name
    tmp_zip.close()

    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as zipf:
            # 1) Datenbank sichern
            if DATABASE_URL:
                conn = get_db()
                try:
                    cur = conn.cursor()
                    for table in ("users", "kosten", "startguthaben", "vorschuesse"):
                        try:
                            cur.execute(f"SELECT * FROM {table}")
                            rows = cur.fetchall() or []
                            write_table_csv(zipf, table, rows)
                        except Exception as table_error:
                            app.logger.exception("Backup-Export für Tabelle %s fehlgeschlagen", table)
                            zipf.writestr(
                                f"database/{table}_ERROR.txt",
                                f"Tabelle {table} konnte nicht exportiert werden: {table_error}".encode("utf-8"),
                            )
                finally:
                    conn.close()
            else:
                if os.path.exists(DB_PATH):
                    # SQLite wird über die Backup-API kopiert. Das ist stabiler als die laufende DB-Datei
                    # direkt zu zippen, besonders wenn WAL-Dateien aktiv sind.
                    snapshot = tempfile.NamedTemporaryFile(prefix="reisekosten_db_snapshot_", suffix=".db", delete=False)
                    snapshot_path = snapshot.name
                    snapshot.close()
                    try:
                        source_conn = sqlite3.connect(DB_PATH)
                        target_conn = sqlite3.connect(snapshot_path)
                        try:
                            source_conn.backup(target_conn)
                        finally:
                            target_conn.close()
                            source_conn.close()
                        zipf.write(snapshot_path, arcname="database/reisekosten.db")
                    finally:
                        try:
                            os.remove(snapshot_path)
                        except OSError:
                            pass
                else:
                    zipf.writestr(
                        "database/INFO.txt",
                        f"SQLite-Datenbank wurde nicht gefunden: {DB_PATH}".encode("utf-8"),
                    )

            # 2) Belege/Uploads sichern
            upload_dir = app.config["UPLOAD_FOLDER"]
            if os.path.isdir(upload_dir):
                for root, _, files in os.walk(upload_dir):
                    for filename in files:
                        full_path = os.path.join(root, filename)
                        if not os.path.isfile(full_path) or os.path.islink(full_path):
                            continue
                        rel_path = os.path.relpath(full_path, upload_dir).replace(os.sep, "/")
                        zipf.write(full_path, arcname=f"uploads/{rel_path}")
            else:
                zipf.writestr("uploads/INFO.txt", b"Kein Upload-Ordner gefunden.")

        response = send_file(
            zip_path,
            as_attachment=True,
            download_name=zip_name,
            mimetype="application/zip",
            max_age=0,
        )
        # Nach dem Senden entfernt Flask die temporäre Datei über diesen Callback.
        response.call_on_close(lambda: os.path.exists(zip_path) and os.remove(zip_path))
        return response

    except Exception as e:
        app.logger.exception("Backup konnte nicht erstellt werden")
        try:
            if os.path.exists(zip_path):
                os.remove(zip_path)
        except OSError:
            pass
        flash(f"Backup konnte nicht erstellt werden: {e}", "err")
        return redirect(url_for("dashboard"))


# =========================
# Belege-Restore-Hilfsfunktionen
# =========================
RESTORE_ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf", ".heic", ".heif"}


def restore_target_relative_name(zip_member_name):
    """Return the filename/path that should be restored under uploads/.

    Supports all known backup formats:
    1. app backup: uploads/<filename>
    2. browser emergency backup: belege_backup/<filename>
    3. browser ZIP root: <filename>
    """
    name = (zip_member_name or "").replace("\\", "/")
    name = posixpath.normpath(name).lstrip("/")
    if not name or name == "." or name.startswith("../") or "/../" in name:
        return None

    lower_name = name.lower()
    if lower_name.startswith("__macosx/"):
        return None
    if lower_name.endswith("fehlende_bilder.txt") or lower_name.endswith("info.txt"):
        return None

    if name.startswith("uploads/"):
        rel_name = name[len("uploads/"):]
    elif name.startswith("belege_backup/"):
        rel_name = name[len("belege_backup/"):]
    else:
        # Browser-ZIP can also contain files directly in the ZIP root.
        rel_name = posixpath.basename(name)

    rel_name = posixpath.normpath(rel_name).lstrip("/")
    if not rel_name or rel_name == "." or rel_name.startswith("../") or "/../" in rel_name:
        return None

    ext = os.path.splitext(rel_name)[1].lower()
    if ext not in RESTORE_ALLOWED_EXTENSIONS:
        return None

    return rel_name


def safe_upload_restore_path(rel_name):
    """Return an absolute target path inside UPLOAD_FOLDER or None."""
    upload_dir = app.config["UPLOAD_FOLDER"]
    target_path = os.path.abspath(os.path.join(upload_dir, rel_name.replace("/", os.sep)))
    upload_root = os.path.abspath(upload_dir) + os.sep
    if not target_path.startswith(upload_root):
        return None
    return target_path


def de_numbered_restore_name(rel_name):
    """Return filename without emergency-browser numeric prefix, e.g. 12_file.jpg -> file.jpg.

    The first emergency console code saved files as 1_<original>, 2_<original>, ... .
    The database usually expects <original>. Saving this alias keeps the link working.
    """
    directory, basename = posixpath.split(rel_name.replace("\\", "/"))
    match = re.match(r"^\d+_(.+)$", basename)
    if not match:
        return None
    alias_base = match.group(1)
    if not alias_base or os.path.splitext(alias_base)[1].lower() not in RESTORE_ALLOWED_EXTENSIONS:
        return None
    return posixpath.join(directory, alias_base) if directory else alias_base


def save_restored_fileobj(source_fileobj, rel_name):
    """Save one restored upload file and, if needed, also save a de-numbered alias."""
    target_path = safe_upload_restore_path(rel_name)
    if target_path is None:
        raise ValueError("Ungültiger Zielpfad")

    os.makedirs(os.path.dirname(target_path), exist_ok=True)
    existed = os.path.exists(target_path)
    with open(target_path, "wb") as target:
        shutil.copyfileobj(source_fileobj, target, length=1024 * 1024)

    aliases = []
    alias_rel = de_numbered_restore_name(rel_name)
    if alias_rel:
        alias_path = safe_upload_restore_path(alias_rel)
        if alias_path and alias_path != target_path:
            os.makedirs(os.path.dirname(alias_path), exist_ok=True)
            shutil.copyfile(target_path, alias_path)
            aliases.append(alias_rel)

    return {"filename": rel_name, "overwritten": existed, "aliases": aliases}


@app.route("/admin/restore-beleg-file", methods=["POST"])
def restore_beleg_file():
    """Restore a single receipt file.

    Used by the admin page's JavaScript restore mode so large backups do not run
    in one long HTTP request. Each file is uploaded and saved separately.
    """
    if session.get("role") != "buchhaltung":
        return jsonify({"ok": False, "error": "Keine Berechtigung."}), 403

    uploaded_file = request.files.get("file")
    rel_name = restore_target_relative_name(request.form.get("filename") or (uploaded_file.filename if uploaded_file else ""))
    if not uploaded_file or not rel_name:
        return jsonify({"ok": False, "error": "Keine gültige Beleg-Datei."}), 400

    try:
        os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
        result = save_restored_fileobj(uploaded_file.stream, rel_name)
        return jsonify({
            "ok": True,
            "filename": result["filename"],
            "overwritten": result["overwritten"],
            "aliases": result["aliases"],
            "alias_count": len(result["aliases"]),
        })
    except Exception as exc:
        app.logger.exception("Einzelner Beleg konnte nicht wiederhergestellt werden: %s", rel_name)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/delete_account", methods=["POST"])
def delete_account():
    if "user_id" not in session:
        flash("Bitte zuerst einloggen.")
        return redirect(url_for("login"))

    user_id = session["user_id"]

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "DELETE FROM users WHERE id = %s" if DATABASE_URL else
            "DELETE FROM users WHERE id = ?",
            (user_id,)
        )
        conn.commit()
    finally:
        conn.close()

    session.clear()
    flash("Dein Account wurde gelöscht. Du kannst dich neu registrieren.")
    return redirect(url_for("register"))


@app.route("/buchhaltung/reisefuehrer_loeschen/<int:user_id>", methods=["POST"])
def buchhalter_delete_guide(user_id):
    # Das Löschen von Reiseführern durch die Buchhaltung ist bewusst deaktiviert.
    # Die Route bleibt nur als Schutz bestehen, falls noch ein alter Link/Button aufgerufen wird.
    flash("Reiseführer können von der Buchhaltung nicht gelöscht werden.", "error")
    return redirect(url_for("dashboard"))

@app.route("/admin/restore-belege", methods=["POST"])
def restore_belege():
    # 🔐 Nur Buchhaltung
    if session.get("role") != "buchhaltung":
        flash("Keine Berechtigung.", "error")
        return redirect(url_for("dashboard"))

    file = request.files.get("backup_zip")
    if not file or not file.filename.lower().endswith(".zip"):
        flash("Bitte eine gültige ZIP-Datei auswählen.", "err")
        return redirect(url_for("dashboard"))

    upload_dir = app.config["UPLOAD_FOLDER"]
    os.makedirs(upload_dir, exist_ok=True)

    tmp_upload = tempfile.NamedTemporaryFile(prefix="reisekosten_restore_", suffix=".zip", delete=False)
    tmp_upload_path = tmp_upload.name

    try:
        # Upload erst auf Platte schreiben. Das ist stabiler als ZipFile direkt auf FileStorage,
        # vor allem bei größeren Backups auf Render/Gunicorn.
        with tmp_upload:
            shutil.copyfileobj(file.stream, tmp_upload, length=1024 * 1024)

        restored = 0
        skipped = 0
        overwritten = 0
        errors = 0

        with zipfile.ZipFile(tmp_upload_path, "r", allowZip64=True) as zipf:
            # Wichtig: kein zipf.testzip() bei vielen Bildern. Das prüft vorher jede Datei komplett
            # und lässt den Upload in der Oberfläche so wirken, als würde er endlos hängen.
            for member in zipf.infolist():
                rel_name = restore_target_relative_name(member.filename)
                if member.is_dir() or not rel_name:
                    skipped += 1
                    continue

                try:
                    with zipf.open(member) as source:
                        result = save_restored_fileobj(source, rel_name)
                    if result.get("overwritten"):
                        overwritten += 1
                    # Eine de-nummerierte Alias-Datei zählt nicht als eigener Beleg,
                    # erhöht aber die Chance, dass alte Datenbank-Links sofort wieder funktionieren.
                    restored += 1
                except Exception:
                    app.logger.exception("Ein Beleg konnte nicht wiederhergestellt werden: %s", member.filename)
                    errors += 1

        app.logger.info(
            "Beleg-Backup wiederhergestellt: %s Dateien, %s überschrieben, %s übersprungen, %s Fehler",
            restored, overwritten, skipped, errors,
        )
        if restored:
            flash(f"✅ {restored} Beleg-Dateien wiederhergestellt. Überschrieben: {overwritten}. Übersprungen: {skipped}. Fehler: {errors}.", "ok")
        else:
            flash(
                "Es wurden keine Beleg-Dateien wiederhergestellt. Bitte prüfen, ob die ZIP Bilder enthält und ob die Dateien .jpg/.png/.webp/.pdf usw. sind.",
                "err",
            )

    except zipfile.BadZipFile:
        app.logger.exception("Ungültige ZIP-Datei beim Wiederherstellen der Belege")
        flash("Fehler beim Wiederherstellen: Die Datei ist keine gültige ZIP-Datei oder ist beschädigt.", "err")
    except Exception as e:
        app.logger.exception("Fehler beim Wiederherstellen der Belege")
        flash(f"Fehler beim Wiederherstellen: {e}", "err")
    finally:
        try:
            os.remove(tmp_upload_path)
        except OSError:
            pass

    return redirect(url_for("dashboard"))



